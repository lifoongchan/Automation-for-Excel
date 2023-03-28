import openpyxl as xl
from openpyxl.chart import BarChart, Reference

wb = xl.load_workbook("transactions.xlsx")
sheet = wb["Sheet1"]

#auto-calculation
for row in range(2, sheet.max_row + 1): #row starts from second row
    cell = sheet.cell(row, 3) #cells start from third column
    corrected_price = float(cell.value) * 0.9
    corrected_price_cell = sheet.cell(row, 4)
    corrected_price_cell.value = corrected_price #insert values in the targeted locations

#for chart/bar - referred to the column and row
values = Reference(sheet,
          min_row=2,
          max_row=sheet.max_row,
          min_col=4,
          max_col=4)

#make chart
chart = BarChart()
chart.add_data(values)
sheet.add_chart(chart, "e2")

#save as
wb.save("transactions2.xlsx")
